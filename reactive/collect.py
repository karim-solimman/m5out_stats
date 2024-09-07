import os
import openpyxl

ws_header = ["#", "routing algorithm", "synthetic traffic", "injection rate", "sim cycles",
                  "average_flit_latency", "average_flit_network_latency", "flit_queuing_latency",
                    "average_flit_vnet_latency", "average_hops", "average_packet_latency", "average_packet_network_latency",
                    "average_packet_queueing_latency", "average_packet_vnet_latency", "average_link_utilization",
                    "average_vc_load", "ext_in_link_utilization", "ext_out_link_utilization",
                    "Flits_inject", "Flits_received", "int_link_utilization", "Pkt_inject", "Pkt_received", 
                    "Flits_delivery_perc", "Pkt_delivery_perc", "throughput", "receiption_rate",]


syn_traffics = ["triba27_uniform_random", "triba27_bit_reverse", "triba27_transpose",
                 "triba27_tornado", "triba27_bit_complement", "triba27_shuffle"]

inj_rates = []


if __name__ == '__main__':

    print("Preparing data...")

    # data extraction
    extension = '.xlsx'
    extracted_data = "average_packet_latency"
    syn_traffic = "triba27_bit_reverse"
    
    # select xlsx files [not throughput files]
    all_files = os.listdir("./")
    xlsx_files = []
    for file in all_files:
        if file.endswith(extension) and not "throughput" in file and not "full-system" in file:
            xlsx_files.append(file)
    xlsx_files.sort()

    print("Creating output xlsx file...")
    # the output xlsx file
    output_header = ['#', 'injection_rate']
    wb = openpyxl.Workbook()
    ws = wb.active

    for file in xlsx_files:
        routing_algorithm = file.split('.')[0]
        output_header.append(routing_algorithm)
    ws.append(output_header)


    # collecting data
    print("Collecting data...")
    output_data = dict()
    for i in range(5, 205, 5):
        inj_rates.append(i/1000)
        output_data[str(i/1000)] = dict()
        for file in xlsx_files:
            routing_algorithm = file.split('.')[0]
            output_data[str(i/1000)][routing_algorithm] = 0.0
    
    # prepare the data in dictionary format
    for file in xlsx_files:
        routing_algorithm = file.split('.')[0]
        tmp_wb = openpyxl.load_workbook(file)
        tmp_wc = tmp_wb['Sheet']
        rows = tmp_wc.max_row
        data_column = ws_header.index(extracted_data)
        injection_rate_column = ws_header.index('injection rate')
        for i in range(rows + 1):
            if tmp_wc.cell(row= i + 1, column = 3).value == syn_traffic:
                inj_rate = tmp_wc.cell(row= i + 1, column=injection_rate_column + 1).value
                collected_data = tmp_wc.cell(row=i + 1, column=data_column + 1).value
                output_data[str(inj_rate)][routing_algorithm] = collected_data
        tmp_wb.close()

    
    # Write the data to the file
    print("Writing data...")
    for index, (inj, values) in enumerate(output_data.items()):
        for algo, data in values.items():
            row = inj_rates.index(float(inj)) + 2
            column = output_header.index(algo) + 1
            ws.cell(row=row, column=column).value = data
            ws.cell(row=row, column=2).value = float(inj)
            ws.cell(row=row, column=1).value = index + 1
    wb.save(extracted_data + "_" + syn_traffic + ".xlsx")

    print(f"{syn_traffic, extracted_data} | Generated sucessfully...")