import openpyxl

routing_algorithms = [
    "ddra",
    "dm4t",
    "dm4t-randomEqual",
    "dm4t-randomEqual-modified",
    "full-adaptive",
    "full-adaptive-randomEqual",
    "semi-adaptive",
    "semi-adaptive-randomEqual",
    "semi-adaptive-randomEqual-modified",
]

synthetic_traffics = [
    "uniform_random",
    "bit_reverse",
    "bit_complement",
    "shuffle",
    "transpose",
    "tornado"
]

if __name__ == "__main__":
    synthetic_traffic = "transpose"
    target_data = "routers_sw_dynamic"
    vc_4_depth_enable = False

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    
    output_header = ["#", "injection_rate"] + routing_algorithms
    output_data = {key:[] for key in routing_algorithms}
    output_ws.append(output_header)
    
    for index, routing_algo in enumerate(routing_algorithms):
        if vc_4_depth_enable:
            excel_file = f"power_vcDepth-4-throughput-{routing_algo}.xlsx"
        else:
            excel_file = f"power_throughput-{routing_algo}.xlsx"
        wb = openpyxl.load_workbook(excel_file)
        ws = wb["Sheet"]
        data_column = -1

        for i in range(1, ws.max_column + 1):
            if target_data in ws.cell(row=1, column=i).value:
                data_column = i
                break

        assert data_column != -1, "can't find the column data in the excel sheet"
        
        for i in range(1, ws.max_row + 1):
            if synthetic_traffic in ws.cell(row=i, column=3).value:
                output_data[routing_algo].append(ws.cell(row=i, column=data_column).value)

        print(excel_file, "Done...")
        wb.close()

    injection_rates = [i/100 for i in range(5, 105, 5)]
    for index, inj_rate in enumerate(injection_rates):
        data = [index + 1 , inj_rate]
        for rout_algo in routing_algorithms:
            data.append(output_data[rout_algo][index])
        output_ws.append(data)
    
    if vc_4_depth_enable:
        output_wb.save(f"vcDepth_4-{target_data}_{synthetic_traffic}.xlsx")
    else:
        output_wb.save(f"{target_data}_{synthetic_traffic}.xlsx")
    print("Output_successfully...")

    
    
        

